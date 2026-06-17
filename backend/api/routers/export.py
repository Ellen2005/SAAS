"""
Export Router
=============
Provides export functionality for reports, charts, and data.

Supported formats:
  - PDF (via WeasyPrint or browser print)
  - Excel (via openpyxl)
  - CSV (native)
  - PNG (via Playwright or client-side canvas)
  - SVG (native)
"""

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response

from ..core.auth import require_role, resolve_user_id
from ..core.supabase_client import get_supabase
from ..services.chart_service import build_chart_from_rows

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/export", tags=["export"])


def _safe(resp) -> list:
    return resp.data if hasattr(resp, "data") and resp.data else []


# ─── CSV Export ──────────────────────────────────────────────────────────────

@router.get("/csv")
def export_csv(
    table: str = Query(..., description="Table name: kpi_results, anomaly_records, etc."),
    user_id: str = Depends(resolve_user_id),
):
    """Export user's data as CSV."""
    supabase = get_supabase()
    
    # Map table names to allowed tables
    allowed_tables = {
        "kpi_results": "kpi_results",
        "anomaly_records": "anomaly_records",
        "daily_reports": "daily_reports",
        "validation_logs": "validation_logs",
    }
    
    if table not in allowed_tables:
        raise HTTPException(status_code=400, detail=f"Table '{table}' not allowed. Allowed: {list(allowed_tables.keys())}")
    
    rows = _safe(
        supabase.table(allowed_tables[table])
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .limit(1000)
        .execute()
    )
    
    if not rows:
        raise HTTPException(status_code=404, detail="No data found")
    
    # Convert to CSV
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    
    csv_data = output.getvalue()
    filename = f"{table}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return Response(
        content=csv_data,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Excel Export ────────────────────────────────────────────────────────────

@router.get("/excel")
def export_excel(
    table: str = Query(..., description="Table name"),
    user_id: str = Depends(resolve_user_id),
):
    """Export user's data as Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl. Install with: pip install openpyxl")
    
    supabase = get_supabase()
    
    allowed_tables = {
        "kpi_results": "kpi_results",
        "anomaly_records": "anomaly_records",
        "daily_reports": "daily_reports",
    }
    
    if table not in allowed_tables:
        raise HTTPException(status_code=400, detail=f"Table '{table}' not allowed.")
    
    rows = _safe(
        supabase.table(allowed_tables[table])
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .limit(1000)
        .execute()
    )
    
    if not rows:
        raise HTTPException(status_code=404, detail="No data found")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table
    
    # Headers
    headers = list(rows[0].keys())
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row in rows:
        ws.append(list(row.values()))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{table}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Chart Export (PNG via SVG conversion) ───────────────────────────────────

@router.get("/chart/png")
def export_chart_png(
    chart_type: str = Query("bar", description="Chart type: bar, line, pie, etc."),
    data: str = Query(..., description="JSON-encoded data array"),
    title: str = Query("Chart", description="Chart title"),
    user_id: str = Depends(resolve_user_id),
):
    """Export a chart as PNG (via SVG conversion)."""
    try:
        import json
        import base64
        from urllib.parse import quote
        
        chart_data = json.loads(data)
        if not chart_data:
            raise HTTPException(status_code=400, detail="No data provided")
        
        # Build chart spec
        columns = list(chart_data[0].keys())
        spec = build_chart_from_rows(chart_data, columns, chart_type=chart_type, title=title)
        
        if not spec:
            raise HTTPException(status_code=400, detail="Could not generate chart")
        
        # Use QuickChart for PNG export
        chart_config = {
            "type": spec["type"],
            "data": {
                "labels": [d.get("label", "") for d in spec["data"]],
                "datasets": [{
                    "label": title,
                    "data": [d.get("value", 0) for d in spec["data"]],
                    "backgroundColor": spec.get("colors", [])[:len(spec["data"])],
                }]
            },
            "options": {
                "title": {"display": True, "text": title},
                "legend": {"position": "bottom"},
            }
        }
        
        # For simplicity, return SVG (PNG requires Playwright)
        svg_url = f"https://quickchart.io/chart?c={quote(json.dumps(chart_config))}&w=800&h=400&bkg=white&format=svg"
        
        return {"svg_url": svg_url, "chart_spec": spec}
        
    except Exception as e:
        logger.error(f"Chart export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Report Export (HTML to PDF) ─────────────────────────────────────────────

@router.get("/report/pdf")
def export_report_pdf(
    report_type: str = Query(..., description="Report type: dg, board, regional"),
    context: dict = Depends(require_role(["manager", "admin"])),
):
    """Export executive report as PDF."""
    from ..services.executive_report_service import generate_pdf_report
    
    supabase = get_supabase()
    user_id = context["user_id"]
    
    kpi_rows = _safe(
        supabase.table("kpi_results")
        .select("*")
        .eq("user_id", user_id)
        .order("recorded_at", desc=True)
        .limit(20)
        .execute()
    )
    kpis = [{"name": r.get("kpi_name"), "value": float(r.get("value", 0)), "change_pct": float(r.get("dod_pct", 0)) if r.get("dod_pct") else None, "status": r.get("status", "NORMAL")} for r in kpi_rows]
    
    anomaly_rows = _safe(
        supabase.table("anomaly_records")
        .select("*")
        .eq("user_id", user_id)
        .order("detected_at", desc=True)
        .limit(20)
        .execute()
    )
    anomalies = [{"kpi_name": r.get("kpi_name"), "severity": r.get("severity", "WARNING"), "deviation": float(r.get("deviation", 0)), "context": r.get("context", {})} for r in anomaly_rows]
    
    report_period = datetime.now().strftime("%B %Y")
    
    if report_type == "dg":
        pdf_bytes, filename, _ = generate_pdf_report(
            "dg",
            company_name="CNPS",
            report_period=report_period,
            kpis=kpis,
            anomalies=anomalies,
            regional_data=[],
            department_performance=[],
            recommendations=[],
            risks=[],
            executive_summary="",
        )
    elif report_type == "board":
        pdf_bytes, filename, _ = generate_pdf_report(
            "board",
            company_name="CNPS",
            report_period=report_period,
            kpis=kpis,
            strategic_objectives=[],
            financial_summary="",
        )
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/chart/svg")
def export_chart_svg(
    chart_type: str = Query("bar", description="Chart type"),
    data: str = Query(..., description="JSON-encoded data array"),
    title: str = Query("Chart", description="Chart title"),
    user_id: str = Depends(resolve_user_id),
):
    """Export chart as SVG."""
    try:
        import json
        from urllib.parse import quote
        
        chart_data = json.loads(data)
        if not chart_data:
            raise HTTPException(status_code=400, detail="No data provided")
        
        columns = list(chart_data[0].keys())
        spec = build_chart_from_rows(chart_data, columns, chart_type=chart_type, title=title)
        
        if not spec:
            raise HTTPException(status_code=400, detail="Could not generate chart")
        
        chart_config = {
            "type": spec["type"],
            "data": {
                "labels": [d.get("label", "") for d in spec["data"]],
                "datasets": [{
                    "label": title,
                    "data": [d.get("value", 0) for d in spec["data"]],
                    "backgroundColor": spec.get("colors", [])[:len(spec["data"])],
                }]
            },
            "options": {
                "title": {"display": True, "text": title},
                "legend": {"position": "bottom"},
            }
        }
        
        svg_url = f"https://quickchart.io/chart?c={quote(json.dumps(chart_config))}&w=800&h=400&bkg=white&format=svg"
        
        return {"svg_url": svg_url, "chart_spec": spec}
    except Exception as e:
        logger.error(f"SVG export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
