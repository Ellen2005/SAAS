"""Export services for KPIs and reports in multiple formats."""
import io
import csv
import logging
from datetime import datetime, timezone
from typing import Optional

logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export will be disabled.")


def export_kpis_csv(user_id: str, supabase) -> bytes:
    """Export KPIs as CSV."""
    try:
        rows = (
            supabase.table("kpi_results")
            .select("*")
            .eq("user_id", user_id)
            .order("recorded_at", desc=True)
            .limit(1000)
            .execute()
        )
        kpis = rows.data if hasattr(rows, "data") and rows.data else []
        
        if not kpis:
            return b"No KPI data available"
        
        # Create CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "kpi_name", "value", "dod_pct", "wow_pct", "avg_7d", "status", "recorded_at"
        ])
        writer.writeheader()
        for kpi in kpis:
            writer.writerow({
                "kpi_name": kpi.get("kpi_name", ""),
                "value": kpi.get("value", 0),
                "dod_pct": kpi.get("dod_pct", ""),
                "wow_pct": kpi.get("wow_pct", ""),
                "avg_7d": kpi.get("avg_7d", ""),
                "status": kpi.get("status", ""),
                "recorded_at": kpi.get("recorded_at", ""),
            })
        
        return output.getvalue().encode("utf-8")
    except Exception as e:
        logger.error(f"CSV export error: {e}")
        raise


def export_kpis_excel(user_id: str, supabase) -> bytes:
    """Export KPIs as Excel with formatting."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
    
    try:
        rows = (
            supabase.table("kpi_results")
            .select("*")
            .eq("user_id", user_id)
            .order("recorded_at", desc=True)
            .limit(1000)
            .execute()
        )
        kpis = rows.data if hasattr(rows, "data") and rows.data else []
        
        if not kpis:
            raise ValueError("No KPI data available")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KPIs"
        
        # Headers with formatting
        headers = ["KPI Name", "Value", "DoD %", "WoW %", "Avg 7D", "Status", "Recorded At"]
        header_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, kpi in enumerate(kpis, 2):
            ws.cell(row=row_num, column=1, value=kpi.get("kpi_name", ""))
            ws.cell(row=row_num, column=2, value=kpi.get("value", 0))
            ws.cell(row=row_num, column=3, value=kpi.get("dod_pct", ""))
            ws.cell(row=row_num, column=4, value=kpi.get("wow_pct", ""))
            ws.cell(row=row_num, column=5, value=kpi.get("avg_7d", ""))
            ws.cell(row=row_num, column=6, value=kpi.get("status", ""))
            ws.cell(row=row_num, column=7, value=kpi.get("recorded_at", ""))
            
            # Color code status
            status_cell = ws.cell(row=row_num, column=6)
            status = kpi.get("status", "").upper()
            if status == "CRITICAL":
                status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                status_cell.font = Font(color="CC0000", bold=True)
            elif status == "WARNING":
                status_cell.fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
                status_cell.font = Font(color="B35900", bold=True)
            elif status == "NORMAL":
                status_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                status_cell.font = Font(color="006600", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise


def export_report_as_excel(report: dict, supabase) -> bytes:
    """Export a report as Excel with narrative and KPIs."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
    
    try:
        wb = Workbook()
        
        # Sheet 1: Narrative
        ws_narrative = wb.active
        ws_narrative.title = "Report Narrative"
        
        narrative = report.get("narrative", "No narrative available")
        ws_narrative.cell(row=1, column=1, value="Report Narrative")
        ws_narrative.cell(row=1, column=1).font = Font(bold=True, size=14, color="1a3a5c")
        ws_narrative.cell(row=3, column=1, value=f"Report Date: {report.get('report_date', 'N/A')}")
        ws_narrative.cell(row=4, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Split narrative into paragraphs
        paragraphs = narrative.split('\n\n')
        current_row = 6
        for para in paragraphs:
            if para.strip():
                cell = ws_narrative.cell(row=current_row, column=1, value=para.strip())
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws_narrative.row_dimensions[current_row].height = max(30, len(para) // 100 * 15)
                current_row += 1
        
        ws_narrative.column_dimensions['A'].width = 100
        
        # Sheet 2: KPIs
        ws_kpis = wb.create_sheet("KPIs")
        user_id = report.get("user_id")
        
        if user_id:
            kpi_rows = (
                supabase.table("kpi_results")
                .select("*")
                .eq("user_id", user_id)
                .order("recorded_at", desc=True)
                .limit(100)
                .execute()
            )
            kpis = kpi_rows.data if hasattr(kpi_rows, "data") and kpi_rows.data else []
            
            if kpis:
                headers = ["KPI Name", "Value", "DoD %", "WoW %", "Status", "Date"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws_kpis.cell(row=1, column=col_num, value=header)
                    cell.fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                for row_num, kpi in enumerate(kpis, 2):
                    ws_kpis.cell(row=row_num, column=1, value=kpi.get("kpi_name", ""))
                    ws_kpis.cell(row=row_num, column=2, value=kpi.get("value", 0))
                    ws_kpis.cell(row=row_num, column=3, value=kpi.get("dod_pct", ""))
                    ws_kpis.cell(row=row_num, column=4, value=kpi.get("wow_pct", ""))
                    ws_kpis.cell(row=row_num, column=5, value=kpi.get("status", ""))
                    ws_kpis.cell(row=row_num, column=6, value=kpi.get("recorded_at", ""))
                
                for column in ws_kpis.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws_kpis.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
        
    except Exception as e:
        logger.error(f"Report Excel export error: {e}")
        raise